## csv_practice.py

import pandas as pd
import openpyxl

# df = pd.DataFrame([[11, 21, 31], [12, 22, 32], [31, 32, 33]],
#                 index=['one', 'two', 'three'], columns=['a', 'b', 'c'])


# print(df)
#         a   b   c
# one    11  21  31
# two    12  22  32
# three  31  32  33

# df2 = df[['a', 'c']]
# print(df2)

l = ['aaa', 'bbb', 'ccc']
m = ['''
     aaa
     bbb
     ccc''']
'''
worksheet.cell('A1').style.alignment.wrap_text = True

worksheet.cell('A1').value = "Line 1\nLine 2\nLine 3"

workbook.save('wrap_text1.xlsx')
'''

from openpyxl.styles import Alignment
ws['A1'].alignment = Alignment(wrap_text=True)

s = '\n'.join(l)
#print(s)
# aaa
# bbb
# ccc
'''
cell_format = workbook.add_format({'bold': True, 'font_color': 'red'})
worksheet.write('A1', 'Cell A1', cell_format)

wrap_format = workbook.add_format({'text_wrap': True})

# Later...
cell_format.set_font_color('green')
worksheet.write('D1', 'Cell B1', cell_format)
'''

df3 = pd.DataFrame([['Load data', 'Ryan', s]], 
     index=['script 1'], columns=['Summary', 'Reporter', 'Description'])

# Pandas
with pd.ExcelWriter('concat_writer.xlsx', engine='openpyxl') as writer:
     # df.to_excel(writer, sheet_name='sheet1')
     # df2.to_excel(writer, sheet_name='sheet2')
     df3.to_excel(writer, sheet_name='sheet3')

#ExcelWriter

wrap_format = workbook.add_format({'text_wrap': True})

# still need
# Workbook
# Worksheet
